import numpy as np
import pandas as pd
from openpyxl.styles import Alignment

# Чтение данных из Excel-файла, пропуская строки перед нужной
df = pd.read_excel('C:/Users/Misha/Desktop/Вюрт/31102024 28102024 ВюртБел.xls', header=None)

# Извлечение названий столбцов из четвёртой строки (индекс 3)
df.columns = df.iloc[3]

# Удаление строк с заголовками
df = df[4:]

# Удаление лишних пробелов в названиях столбцов
df.columns = df.columns.str.strip()

# Выбор необходимых столбцов
columns_to_keep = ['ТМЦ', 'ЕдиницаИзмерения', 'Количество', 'ПерваяЦена', 'СуммаСНДС','Остаток']
df_filtered = df[columns_to_keep]

# Сохранение отфильтрованного DataFrame в новый Excel-файл,
# начиная с 2-й строки (индекс 1) и 2-го столбца (индекс 1)
df_filtered.to_excel('C:/Users/Misha/Desktop/Вюрт/2.xlsx', index=False, header=False, startrow=1, startcol=1)

print("Данные успешно сохранены, начиная с 2-й строки и 2-го столбца.")






# Чтение данных из Excel-файла и создание DataFrame
df = pd.read_excel('C:/Users/Misha/Desktop/Вюрт/2.xlsx')

# Номер колонки, для которой нужно присвоить числовой индекс (нумерация начинается с 0)
column_number = 1

# Присвоение числового индекса на основе текста в указанной колонке
df = df.set_index(df.columns[column_number])

# Сортировка DataFrame по индексу
df_sorted = df.sort_values(by=df.index.name)

# Агрегирование строк с одинаковым индексом
aggregation_dict = {'Unnamed: 2': 'first', 'Unnamed: 3': 'sum', 'Unnamed: 4': 'sum', 'Unnamed: 5': 'sum','Unnamed: 6': 'first'}
df_aggregated = df_sorted.groupby(level=0).agg(aggregation_dict)

# Вычисление результата деления значений столбца "Unnamed: 5" на значения столбца "Unnamed: 3"
df_aggregated["Unnamed: 4"] = df_aggregated["Unnamed: 5"] / df_aggregated["Unnamed: 3"]

# Расчет доли каждого товара
df_aggregated['Доля'] = df_aggregated['Unnamed: 5'] / df_aggregated['Unnamed: 5'].sum()

# Расчет кумулятивной доли
df_aggregated['Аккум.доля'] = df_aggregated['Доля'].cumsum()

# Классификация товаров по ABC-анализу
df_aggregated["Категория (ABC)"] = ' '
df_aggregated.loc[df_aggregated['Аккум.доля'] < 0.6, 'Категория (ABC)'] = 'A'
df_aggregated.loc[(df_aggregated['Аккум.доля'] >= 0.6) & (df_aggregated['Аккум.доля'] < 0.75), 'Категория (ABC)'] = 'B'
df_aggregated.loc[df_aggregated['Аккум.доля'] >= 0.75, 'Категория (ABC)'] = 'C'

# Выбор столбца с "Расходом годовым"
df_aggregated['Расход годовой'] = df_aggregated['Unnamed: 3']

# Проверка на нечисловые данные
if not pd.to_numeric(df_aggregated['Расход годовой'], errors='coerce').notnull().all():
    print("Столбец 'Расход годовой' содержит нечисловые данные. Обработайте их перед выполнением анализа.")
    exit()

# Проверка на нулевые значения
if df_aggregated['Расход годовой'].eq(0).sum() > 0:
    print("Столбец 'Расход годовой' содержит нулевые значения. Обработайте их перед выполнением анализа.")
    exit()

# Расчет стандартного отклонения
std = df_aggregated['Расход годовой'].std(ddof=0)

# Расчет среднего значения
mean = df_aggregated['Расход годовой'].mean()

# Проверка на деление на ноль
if mean == 0:
    print("Среднее значение равно нулю. Коэффициент вариации не может быть рассчитан.")
    exit()

# Расчет коэффициента вариации
cv = (std / mean) * 100

# Классификация товаров по XYZ-анализу
df_aggregated['Категория (XYZ)'] = ' '
df_aggregated.loc[df_aggregated['Расход годовой'] <= 0.25 * mean, 'Категория (XYZ)'] = 'Z'
df_aggregated.loc[(df_aggregated['Расход годовой'] > 0.25 * mean) & (
            df_aggregated['Расход годовой'] <= 0.5 * mean), 'Категория (XYZ)'] = 'Y'
df_aggregated.loc[df_aggregated['Расход годовой'] > 0.5 * mean, 'Категория (XYZ)'] = 'X'

# Изменение названий столбцов
df_aggregated.rename(columns={'Unnamed: 2': 'Единицы измерения',
                              'Unnamed: 3': 'Количество',
                              'Unnamed: 4': 'Цена за шт без НДС в среднем',
                              'Unnamed: 5': 'Сумма без НДС',
                              'Unnamed: 6': 'Остаток'
                              }, inplace=True)

# Создание нового DataFrame с наименованиями
df_with_names = df_aggregated.copy()
df_with_names.insert(0, 'Наименование', df_aggregated.index)

# Добавление нового столбца "СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ"
df_with_names['СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ'] = ''

# Заполнение столбца в зависимости от условий
condition_fixed_interval_time = (
        (df_with_names["Категория (ABC)"].isin(['A', 'B'])) &
        (df_with_names["Категория (XYZ)"] == 'X')
)

condition_fixed_order_size = (
        (df_with_names["Категория (ABC)"].isin(['A', 'B'])) &
        (df_with_names["Категория (XYZ)"] == 'Y')
)

condition_fixed_replenishment_periodicity = (
        (df_with_names["Категория (ABC)"].isin(['A', 'B'])) &
        (df_with_names["Категория (XYZ)"] == 'Z')
)

condition_minimum_maximum_system = (
        (df_with_names["Категория (ABC)"] == 'C') &
        (df_with_names["Категория (XYZ)"].isin(['X', 'Y', 'Z']))
)

# Заполнение значений в зависимости от условий
df_with_names.loc[
    condition_fixed_interval_time, 'СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ'] = "Система с фиксированным интервалом времени между заказами"
df_with_names.loc[condition_fixed_order_size, 'СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ'] = "Система с фиксированным размером заказа"
df_with_names.loc[
    condition_fixed_replenishment_periodicity, 'СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ'] = "Система с установленной периодичностью пополнения запасов до установленного уровня"
df_with_names.loc[
    condition_minimum_maximum_system, 'СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ'] = "Система управления запасами «минимум–максимум"

# Рассчитываем МЖЗ, ГЗ, ПУ qо и I для  систем
df_with_names['Максимальный желательный уровень запасов (МЖЗ)'] = 0.0
df_with_names['Гарантийный уровень запасов (ГЗ)'] = 0.0
df_with_names['Пороговый уровень запасов (ПУ)'] = 0.0
df_with_names['Оптимальный размер заказа (qо)'] = 0.0
df_with_names['Интервал времени между заказами (I)'] = 0.0

# Рабочие дни в году
working_days_in_year = 252

# Время выполнения заказа (в рабочих днях)
order_lead_time = 5

# Время задержки (в рабочих днях)
delay_time = 1









                                         # Условия для строк с фиксированным интервалом времени
condition_fixed_interval = df_with_names["СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ"] == "Система с фиксированным интервалом времени между заказами"

# Расчеты для строк с фиксированным интервалом времени


# Максимальный желательный уровень запасов (МЖЗ)
df_with_names.loc[condition_fixed_interval, 'Максимальный желательный уровень запасов (МЖЗ)'] = (
    (df_with_names['Количество'] / working_days_in_year) * (order_lead_time + delay_time)).astype('float64')

# Гарантийный уровень запасов (ГЗ)
df_with_names.loc[condition_fixed_interval, 'Гарантийный уровень запасов (ГЗ)'] = (
    (df_with_names['Количество'] / working_days_in_year) * delay_time).astype('float64')




#Код для заполнения "Интервал времени между заказами (I)"
df_with_names.loc[condition_fixed_interval, 'Интервал времени между заказами (I)'] = working_days_in_year/((df_with_names.loc[condition_fixed_interval, 'Количество'])/(np.sqrt(
    (2 * df_with_names.loc[condition_fixed_interval, 'Количество']))).astype('float64'))

# Расчеты для строк с фиксированным размером заказа
df_with_names.loc[condition_fixed_interval, 'Максимальный желательный уровень запасов (МЖЗ)'] = df_with_names.loc[condition_fixed_interval, 'Гарантийный уровень запасов (ГЗ)']+(df_with_names.loc[condition_fixed_interval, 'Интервал времени между заказами (I)']*(df_with_names['Количество'] / working_days_in_year))

# Оптимальный размер заказа (qо)
df_with_names.loc[condition_fixed_interval, 'Оптимальный размер заказа (qо)'] = df_with_names.loc[condition_fixed_interval, 'Максимальный желательный уровень запасов (МЖЗ)']-df_with_names['Остаток']+((df_with_names['Количество'] / working_days_in_year) * order_lead_time)




                                            # Условия для строк с фиксированным размером заказа

# Условия для строк с фиксированным размером заказа
condition_fixed_order_size = df_with_names["СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ"] == "Система с фиксированным размером заказа"

# Расчеты для строк с фиксированным размером заказа


# Расчет ГЗ
df_with_names.loc[condition_fixed_order_size, 'Гарантийный уровень запасов (ГЗ)'] = (
    (df_with_names['Количество'] / working_days_in_year) * delay_time).astype('float64')

# Пороговый уровень запасов (ПУ)
df_with_names.loc[condition_fixed_order_size, 'Пороговый уровень запасов (ПУ)'] = (
    df_with_names.loc[condition_fixed_order_size, 'Гарантийный уровень запасов (ГЗ)'] +
    (df_with_names['Количество'] / working_days_in_year) * order_lead_time).astype('float64')

# Оптимальный размер заказа (qо)
df_with_names.loc[condition_fixed_order_size, 'Оптимальный размер заказа (qо)'] = np.sqrt(
    (2 * df_with_names.loc[condition_fixed_order_size, 'Количество'])
).astype('float64')

# Расчеты для строк с фиксированным размером заказа
df_with_names.loc[condition_fixed_interval, 'Максимальный желательный уровень запасов (МЖЗ)'] = df_with_names.loc[condition_fixed_interval, 'Гарантийный уровень запасов (ГЗ)']+(df_with_names.loc[condition_fixed_interval, 'Интервал времени между заказами (I)']*(df_with_names['Количество'] / working_days_in_year)).astype('float64')

# Размер заказа (РЗ 1)
df_with_names.loc[condition_fixed_interval, 'Размер заказа (РЗ 1)'] = df_with_names.loc[condition_fixed_interval, 'Максимальный желательный уровень запасов (МЖЗ)']-df_with_names['Остаток']+((df_with_names['Количество'] / working_days_in_year) * order_lead_time).astype('float64')

#Код для заполнения "Интервал времени между заказами (I)"
df_with_names.loc[condition_fixed_order_size, 'Интервал времени между заказами (I)'] = 0


                       # Условия для строк с установленной периодичностью пополнения запасов
# Условия для строк с установленной периодичностью пополнения запасов
condition_fixed_replenishment_periodicity = df_with_names["СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ"] == "Система с установленной периодичностью пополнения запасов до установленного уровня"


#Гарантийный запас (ГЗ) 2
df_with_names.loc[condition_fixed_replenishment_periodicity, 'Гарантийный уровень запасов (ГЗ)'] = ((df_with_names['Количество'] / working_days_in_year) * order_lead_time).astype('float64')

# Пороговый уровень запасов (ПУ) 3
df_with_names.loc[condition_fixed_replenishment_periodicity, 'Пороговый уровень запасов (ПУ)'] =df_with_names.loc[condition_fixed_replenishment_periodicity, 'Гарантийный уровень запасов (ГЗ)'] +(df_with_names['Количество'] / (working_days_in_year) * order_lead_time).astype('float64')

#Ожидаемое потребление товара на складе (ОП) 4
df_with_names.loc[condition_fixed_replenishment_periodicity, 'Ожидаемое потребление товара(ОП)'] =(df_with_names['Количество'] / (working_days_in_year) * order_lead_time).astype('float64')

# Оптимальный размер заказа (qо) 5.1
df_with_names.loc[condition_fixed_replenishment_periodicity, 'Оптимальный размер заказа (qо)'] =  np.sqrt(2 * df_with_names.loc[condition_fixed_replenishment_periodicity, 'Количество'] / working_days_in_year).astype('float64')

# Интервал времени между заказами (I) 5.2
df_with_names.loc[condition_fixed_replenishment_periodicity, 'Интервал времени между заказами (I)'] = ((working_days_in_year * df_with_names.loc[condition_fixed_replenishment_periodicity, 'Оптимальный размер заказа (qо)']) / df_with_names['Количество']).astype('float64')


# Максимальный желательный уровень запасов (МЖЗ) 5
df_with_names.loc[condition_fixed_replenishment_periodicity, 'Максимальный желательный уровень запасов (МЖЗ)'] = df_with_names.loc[condition_fixed_replenishment_periodicity, 'Гарантийный уровень запасов (ГЗ)']+df_with_names.loc[condition_fixed_replenishment_periodicity,'Интервал времени между заказами (I)']*df_with_names.loc[condition_fixed_replenishment_periodicity, 'Ожидаемое потребление товара(ОП)'].astype('float64')


# Размер заказа (РЗ 1) 6
df_with_names.loc[condition_fixed_replenishment_periodicity, 'Размер заказа (РЗ 1)'] = df_with_names.loc[condition_fixed_replenishment_periodicity, 'Максимальный желательный уровень запасов (МЖЗ)']-df_with_names['Остаток']+((df_with_names['Количество'] / working_days_in_year) * order_lead_time).astype('float64')


#Дополнительный заказ (РД) 7
df_with_names.loc[condition_fixed_replenishment_periodicity, 'Первый заказ (РД)'] =df_with_names.loc[condition_fixed_replenishment_periodicity, 'Максимальный желательный уровень запасов (МЖЗ)']-df_with_names.loc[condition_fixed_replenishment_periodicity, 'Пороговый уровень запасов (ПУ)']+df_with_names.loc[condition_fixed_replenishment_periodicity, 'Ожидаемое потребление товара(ОП)'].astype('float64')

#Второй заказ (РЗ 2) 8
df_with_names.loc[condition_fixed_replenishment_periodicity, 'Второй заказ (РЗ 2)'] =df_with_names.loc[condition_fixed_replenishment_periodicity, 'Максимальный желательный уровень запасов (МЖЗ)']-df_with_names.loc[condition_fixed_replenishment_periodicity, 'Пороговый уровень запасов (ПУ)']+df_with_names.loc[condition_fixed_replenishment_periodicity, 'Ожидаемое потребление товара(ОП)'].astype('float64')


# Условие для фильтрации строк
condition = df_with_names["СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ"] == "Система управления запасами «минимум–максимум"

# Столбцы, в которых нужно заменить нули
columns_to_replace = ['Максимальный желательный уровень запасов (МЖЗ)',
                       'Гарантийный уровень запасов (ГЗ)',
                       'Пороговый уровень запасов (ПУ)',
                       'Оптимальный размер заказа (qо)',
                       'Интервал времени между заказами (I)']

# Преобразование типов данных в object, чтобы избежать предупреждений
df_with_names[columns_to_replace] = df_with_names[columns_to_replace].astype(object)

# Замена нулей на текст "в ручном режиме" в указанных столбцах
for column in columns_to_replace:
    df_with_names.loc[condition & (df_with_names[column] == 0), column] = "в ручном режиме"

print("Нули были успешно заменены на 'в ручном режиме' в указанных столбцах.")

# Определение порядка сортировки
order = ['AX', 'AY', 'AZ', 'BX', 'BY', 'BZ', 'CX', 'CY', 'CZ']

# Создание нового столбца для хранения порядка
df_with_names['Сортировка'] = df_with_names['Категория (ABC)'] + df_with_names['Категория (XYZ)']

# Применение пользовательского порядка сортировки
df_with_names['Сортировка'] = pd.Categorical(df_with_names['Сортировка'], categories=order, ordered=True)

# Сортировка DataFrame
df_sorted_final = df_with_names.sort_values(by='Сортировка')

# Удаление временного столбца 'Сортировка'
df_sorted_final.drop(columns=['Сортировка'], inplace=True)

# Сохранение нового DataFrame в Excel-файл с использованием openpyxl для настройки ширины столбцов и переноса текста
output_file_path = r'C:/Users/Misha/Desktop/Вюрт/main_AX_AY_AZ_BX_BY_BZ_CX_CY_CZ_analysis_results3.xlsx'
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    df_sorted_final.to_excel(writer, index=False, sheet_name='Analysis Results')

    # Получаем доступ к рабочему листу и устанавливаем ширину столбцов и перенос текста
    workbook = writer.book
    worksheet = writer.sheets['Analysis Results']

    # Установка ширины для всех столбцов по длине заголовков и увеличиваем ширину для определенных столбцов
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = max_length + 2  # Добавляем небольшой отступ к ширине
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Устанавливаем ширину для столбца "Наименование" и "СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ" равной максимальной длине содержимого этих столбцов.

    # Получаем буквы нужных столбцов.
    name_column_letter = worksheet.cell(row=1, column=1).column_letter  # Столбец "Наименование"
    system_column_letter = worksheet.cell(row=1, column=worksheet.max_column).column_letter  # Последний столбец - "СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ"

    # Устанавливаем ширину для "Наименование"
    max_length_name_col = max(len(str(cell.value)) for cell in worksheet[name_column_letter])
    worksheet.column_dimensions[name_column_letter].width = max_length_name_col + 2

    # Устанавливаем ширину для "СИСТЕМА УПРАВЛЕНИЯ ЗАПАСАМИ"
    max_length_system_col = max(len(str(cell.value)) for cell in worksheet[system_column_letter])
    worksheet.column_dimensions[system_column_letter].width = max_length_system_col + 2

    # Применяем перенос текста к ячейкам в этих столбцах
    for col_letter in [name_column_letter, system_column_letter]:
        for cell in worksheet[col_letter]:
            cell.alignment = Alignment(wrap_text=True)

print("Файл успешно сохранен с обновленными настройками ширины столбцов и переносом текста.")